import os # <-- ¡AÑADE ESTA LÍNEA!

import fitz  # PyMuPDF
from PIL import Image

from pdf2image import convert_from_bytes
import pytesseract
import re
from pypdf import PdfReader
import numpy as np
from deskew import determine_skew
from skimage.color import rgb2gray
from skimage.transform import rotate
# nucleo/views.py
import base64
import openpyxl
import cv2
from django.core.paginator import Paginator # <-- ¡AÑADE ESTA LÍNEA!

from django.core.files import File # <-- ¡AÑADE ESTA LÍNEA!
from datetime import datetime
from io import BytesIO
from datetime import date
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, HttpResponseForbidden
from django.db.models import Count, Q
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from django.contrib import messages

from django.conf import settings
from django.contrib.auth.models import User, Group
from django.contrib.staticfiles.storage import staticfiles_storage
from xhtml2pdf import pisa
from openpyxl.utils import get_column_letter
from .models import Pqrs, ArchivoAdjunto, CalidadPeticionario, TipoTramite, ArchivoAdjunto, Seguimiento 
from .forms import PqrsForm, PqrsFilterForm, AbogadoPqrsForm, ArchivoAdjuntoForm, PdfUploadForm, SeguimientoForm, RespuestaTramiteForm
from .forms import TrasladoPqrsForm # Asegúrate de importar el nuevo formulario
from django.core.files.storage import FileSystemStorage

# --- Funciones de Ayuda ---

def es_coordinador(user):
    return user.groups.filter(name='Coordinadores').exists()

def es_coordinador_o_abogado(user):
    """
    Verifica si un usuario pertenece al grupo 'Coordinadores' O 'Abogados'.
    """
    return user.groups.filter(name__in=['Coordinadores', 'Abogados']).exists()
# --- Vistas Principales ---
# nucleo/views.py

def puede_gestionar_respuesta(user):
    """Verifica si un usuario es Coordinador O Abogado."""
    return user.groups.filter(name__in=['Coordinadores', 'Abogados']).exists()
# nucleo/views.py

@login_required
def dashboard(request):
    # 1. Define el queryset base, EXCLUYENDO 'Anulado' desde el principio.
    base_queryset = Pqrs.objects.exclude(estado='Anulado')

    es_coord = es_coordinador(request.user)
    
    # Si el usuario es un abogado, el queryset base para él son solo sus casos.
    if not es_coord and not request.user.is_superuser:
        base_queryset = base_queryset.filter(responsable=request.user)
    
    # 2. CALCULA LOS CONTEOS PARA LAS TARJETAS usando el queryset base.
    #    Esto asegura que el total y los demás contadores son correctos y no incluyen los anulados.
    total_pqrs = base_queryset.count()
    conteo_recibido = base_queryset.filter(estado='Recibido').count()
    conteo_en_tramite = base_queryset.filter(estado='En Trámite').count()
    conteo_resuelto = base_queryset.filter(estado='Resuelto').count()
    lista_abogados = User.objects.filter(groups__name='Abogados')

    # 3. APLICA LOS FILTROS DEL FORMULARIO a una copia del queryset para la tabla.
    queryset_filtrado = base_queryset
    filter_form = PqrsFilterForm(request.GET, user=request.user)

    if filter_form.is_valid():
        q = filter_form.cleaned_data.get('q')
        vigencia = filter_form.cleaned_data.get('vigencia')
        responsable = filter_form.cleaned_data.get('responsable')
        estado = filter_form.cleaned_data.get('estado')
        
        if q:
            queryset_filtrado = queryset_filtrado.filter(Q(radicado__icontains=q) | Q(asunto__icontains=q))
        
        if vigencia:
            queryset_filtrado = queryset_filtrado.filter(fecha_recepcion_inicial__year=vigencia)
        elif not estado: # Solo aplicar filtro de año actual si no se filtra por estado
            current_year = date.today().year
            queryset_filtrado = queryset_filtrado.filter(fecha_recepcion_inicial__year=current_year)

        if responsable and (es_coord or request.user.is_superuser):
            queryset_filtrado = queryset_filtrado.filter(responsable=responsable)
        
        if estado:
            queryset_filtrado = queryset_filtrado.filter(estado=estado)
        pass
    
    # 4. Ordena el resultado final que se mostrará en la tabla.
    lista_pqrs = queryset_filtrado.order_by('-fecha_recepcion_inicial')

    lista_ordenada = queryset_filtrado.order_by('-fecha_recepcion_inicial')

    # --- ¡NUEVO BLOQUE DE PAGINACIÓN! ---
    # 1. Creamos un Paginator con la lista completa de resultados, mostrando 15 por página.
    # --- ¡NUEVO BLOQUE DE PAGINACIÓN CON OPCIÓN DE PER_PAGE! ---
    # 1. Obtenemos cuántos registros por página quiere ver el usuario (default: 15)
    per_page = request.GET.get('per_page', 10)

    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 15

    # 2. Creamos el paginador con ese tamaño de página
    paginator = Paginator(lista_ordenada, per_page)

    # 3. Obtenemos el número de página que el usuario quiere ver
    page_number = request.GET.get('page')

    # 4. Obtenemos el objeto Page para esa página
    page_obj = paginator.get_page(page_number)
    # --- FIN DEL BLOQUE ---

    # --- FIN DEL NUEVO BLOQUE --
    # 5. Prepara el contexto final para la plantilla.
    contexto = {
        'total_pqrs': total_pqrs,
        'conteo_recibido': conteo_recibido,
        'conteo_en_tramite': conteo_en_tramite,
        'conteo_resuelto': conteo_resuelto,
        'lista_pqrs': page_obj,   # 👈 ahora la lista paginada
        'filter_form': filter_form,
        'es_coordinador': es_coord,
        'lista_abogados': lista_abogados,
    }
    return render(request, 'nucleo/dashboard.html', contexto)


# nucleo/views.py
# nucleo/views.py
@login_required
@user_passes_test(es_coordinador)
def crear_pqrs_desde_pdf(request):
    # (La configuración de Tesseract y Poppler se queda igual)
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    path_a_poppler = r'C:\poppler\Library\bin'

    if request.method == 'POST':
        form = PdfUploadForm(request.POST, request.FILES)
        if form.is_valid():
            pdf_file = request.FILES['pdf_file']

            # --- ¡NUEVO BLOQUE! Guardar PDF temporalmente ---
            fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'tmp'))
            nombre_archivo_temporal = fs.save(pdf_file.name, pdf_file)
            request.session['pdf_original_temporal'] = nombre_archivo_temporal
            # --- FIN DEL NUEVO BLOQUE ---
            try:
                pdf_file.seek(0) # <-- Pequeño ajuste para la lectura OCR
                # 1. Leer texto con OCR y enderezar imagen
                imagenes = convert_from_bytes(pdf_file.read(), poppler_path=path_a_poppler)
                texto_completo = ""
                for imagen_original in imagenes:
                    imagen_gris = rgb2gray(np.array(imagen_original))
                    angulo = determine_skew(imagen_gris)
                    imagen_corregida_array = rotate(np.array(imagen_original), angulo, resize=True) * 255
                    imagen_corregida_pil = Image.fromarray(imagen_corregida_array.astype(np.uint8))
                    texto_completo += pytesseract.image_to_string(imagen_corregida_pil, lang='spa') + "\n"
                
                # --- INICIO DE DIAGNÓSTICO ---
                print("\n\n===================================")
                print("INICIO DE TEXTO CRUDO EXTRAÍDO CON OCR")
                print("===================================")
                print(texto_completo)
                print("===================================")
                print("FIN DE TEXTO CRUDO EXTRAÍDO CON OCR")
                print("===================================\n\n")
                # --- FIN DE DIAGNÓSTICO ---

                # 2. Extracción de datos básicos
                radicado_extraido, asunto_extraido, peticionario_extraido, email_extraido = "", "", "", ""
                match_radicado = re.search(r"VU\s*(\d+)", texto_completo, re.IGNORECASE)
                if match_radicado: radicado_extraido = "VU-" + match_radicado.group(1).strip()
                
                match_asunto = re.search(r"Cordial saludo,([\s\S]+?)(?=Por lo anterior|Es preciso)", texto_completo, re.IGNORECASE)
                if match_asunto: asunto_extraido = " ".join(match_asunto.group(1).strip().split())
                
                if asunto_extraido:
                    match_peticionario = re.search(r"([A-ZÁÉÍÓÚÑ]{2,}\s[A-ZÁÉÍÓÚÑ\s,]+[A-ZÁÉÍÓÚÑ])", asunto_extraido)
                    if match_peticionario: peticionario_extraido = match_peticionario.group(1).strip().rstrip(',')
                    # --- ¡NUEVO BLOQUE! Extracción de Fecha de Recepción (CON DIAGNÓSTICO) ---
                    fecha_recepcion_extraida = None
                    meses_es = {
                        'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
                        'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12
                    }

                    # Plan A (NUEVA PRIORIDAD): Buscamos la línea "Date:" del correo original.
                    for linea in texto_completo.splitlines():
                        if 'date:' in linea.lower():
                            match_fecha_correo = re.search(r"(\d{1,2})(?:\s+de)?\s+([a-z]{3,})(?:\s+de)?\s+(\d{4})", linea, re.IGNORECASE)
                            if match_fecha_correo:
                                try:
                                    dia = int(match_fecha_correo.group(1))
                                    mes_str = match_fecha_correo.group(2).lower()[:3]
                                    ano = int(match_fecha_correo.group(3))
                                    if mes_str in meses_es:
                                        fecha_recepcion_extraida = date(ano, meses_es[mes_str], dia)
                                        print(f"  [LOG] Fecha encontrada en datos del correo (Prioridad 1): {fecha_recepcion_extraida}")
                                        break 
                                except (ValueError, IndexError):
                                    continue

                    # Plan B: Si no se encontró, buscamos la fecha del encabezado del documento (ej. "Popayán, ...")
                    if not fecha_recepcion_extraida:
                        match_fecha_encabezado = re.search(r"Popayán,\s*(\d{1,2})\s+de\s+([a-zA-Z]+)\s+de\s+(\d{4})", texto_completo, re.IGNORECASE)
                        if match_fecha_encabezado:
                            try:
                                dia = int(match_fecha_encabezado.group(1))
                                mes_str = match_fecha_encabezado.group(2).lower()[:3]
                                ano = int(match_fecha_encabezado.group(3))
                                if mes_str in meses_es:
                                    fecha_recepcion_extraida = date(ano, meses_es[mes_str], dia)
                                    print(f"  [LOG] Fecha encontrada en encabezado del PDF (Prioridad 2): {fecha_recepcion_extraida}")
                            except (ValueError, IndexError):
                                pass

                    # Respaldo Final: Si todo lo demás falla, usamos la fecha actual
                    if not fecha_recepcion_extraida:
                        fecha_recepcion_extraida = date.today()
                        print(f"  [LOG] No se encontró ninguna fecha. Usando fecha actual por defecto: {fecha_recepcion_extraida}")

                    # Convertimos a texto para guardarlo en la sesión
                    fecha_para_sesion = fecha_recepcion_extraida.strftime('%Y-%m-%d')

                    # --- FIN DEL NUEVO BLOQUE ---


                # --- ¡NUEVO BLOQUE SEGURO! Añade la fecha de vencimiento si la encuentra ---
                match_vencimiento = re.search(r"(vence el d[ií]a,[\s\S]+?\d{4})", texto_completo, re.IGNORECASE)
                if match_vencimiento:
                    texto_vencimiento = " ".join(match_vencimiento.group(1).strip().split())
                    if asunto_extraido:
                        asunto_extraido += f" (ATENCIÓN: {texto_vencimiento})"
                # --- FIN DEL NUEVO BLOQUE ---

                # 3. Búsqueda de email por capas (Tu versión estable)
                if peticionario_extraido:
                    match_etiqueta = re.search(r"Correo electrónico:\s*(.+)", texto_completo, re.IGNORECASE)
                    if match_etiqueta:
                        email_extraido = match_etiqueta.group(1).strip().splitlines()[0].replace(" ", "")
                    if not email_extraido:
                        patron_email_flexible = r'[\w\.-]+(?:@|Q|Y|\(d|\(M|W)[\w\.-]+'
                        correos_a_ignorar = ['rectoria@', 'quejasreclamos@', 'viceacad@', 'vri@', 'secgral@']
                        nombres_peticionario = peticionario_extraido.lower().split()
                        lineas = texto_completo.splitlines()
                        for i, linea in enumerate(lineas):
                            linea_limpia = linea.strip().lower()
                            if linea_limpia.startswith('de:') and any(nombre in linea_limpia for nombre in nombres_peticionario):
                                contexto_busqueda = "".join(lineas[i:i+4])
                                email_match = re.search(patron_email_flexible, contexto_busqueda)
                                if email_match:
                                    email_encontrado = email_match.group(0)
                                    if not any(ignorado in email_encontrado for ignorado in correos_a_ignorar):
                                        email_extraido = email_encontrado
                                        break
                
                # --- LÓGICA DE CLASIFICACIÓN DE PETICIONARIO ---
                calidad_peticionario_id = None
                calidad_peticionario_detectada = "Externo / Particular" 

                diccionario_calidad = {
                    'Estudiante': ['estudiante', 'alumno', 'judicatura'],
                    'Profesor': ['profesor', 'docente'],
                    'Egresado': ['egresado', 'exalumno'],
                    'Directivo': ['directivo', 'rector', 'vicerrector', 'jefe', 'secretaria general', 'sintraunicol', 'sindicato', 'aspu'],
                    'Funcionario': ['funcionario', 'profesional universitario', 'tecnico administrativo', 'empleado', 'contratista'],
                    'Entidad Gubernamental': ['representante a la camara', 'ministro', 'ministerio', 'consejal', 'senador', 'congresista', 'juzgado', 'tribunal'],
                    'Externo / Particular': ['madre', 'padre', 'representante legal', 'particular']
                }
                
                texto_a_buscar = asunto_extraido.lower() if asunto_extraido else texto_completo.lower()
                
                for calidad, palabras_clave in diccionario_calidad.items():
                    if any(palabra in texto_a_buscar for palabra in palabras_clave):
                        calidad_peticionario_detectada = calidad
                        break 
                
                try:
                    calidad_obj = CalidadPeticionario.objects.get(tipo=calidad_peticionario_detectada)
                    calidad_peticionario_id = calidad_obj.id
                except CalidadPeticionario.DoesNotExist:
                    try:
                        calidad_obj = CalidadPeticionario.objects.get(tipo="Externo / Particular")
                        calidad_peticionario_id = calidad_obj.id
                    except CalidadPeticionario.DoesNotExist:
                        calidad_peticionario_id = None
                # --- ¡NUEVO BLOQUE CORREGIDO! LÓGICA DE CLASIFICACIÓN DE TIPO DE TRÁMITE ---
                tipo_tramite_id = None
                tipo_tramite_detectado = None

                # Aplicamos las reglas con los nombres EXACTOS de tu base de datos
                if calidad_peticionario_detectada == 'Entidad Gubernamental':
                    tipo_tramite_detectado = 'peticiones especiales'  # CORREGIDO
                elif 'queja' in asunto_extraido.lower():
                    tipo_tramite_detectado = 'queja'
                elif 'gratuidad' in texto_completo.lower():
                    tipo_tramite_detectado = 'petición general'  # CORREGIDO
                elif 'documentos' in texto_completo.lower():
                    tipo_tramite_detectado = 'petición de documentos'  # CORREGIDO
                else:
                # --- ¡ESTE ES EL CAMBIO! ---
                # Si ninguna de las condiciones anteriores se cumple, asignamos este por defecto.
                     tipo_tramite_detectado = 'petición general'
                # Si detectamos un tipo de trámite, buscamos su ID en la base de datos
                if tipo_tramite_detectado:
                    try:
                        # Usamos __iexact para ignorar mayúsculas/minúsculas
                        tramite_obj = TipoTramite.objects.get(nombre__iexact=tipo_tramite_detectado)
                        tipo_tramite_id = tramite_obj.id
                        print(f"Tipo de Trámite detectado: '{tipo_tramite_detectado}' (ID: {tipo_tramite_id})")
                    except TipoTramite.DoesNotExist:
                        print(f"Advertencia: Tipo de trámite '{tipo_tramite_detectado}' no encontrado en la base de datos.")
                        tipo_tramite_id = None
                # --- FIN DEL NUEVO BLOQUE ---
                # 4. Limpieza y Reconstrucción final del Email
                if email_extraido:
                    email_extraido = re.sub(r'(Q|Y|\(d|\(M|W)', '@', email_extraido)
                    if '@' not in email_extraido:
                        match_dominio = re.search(r'(unicauca\.edu\.co|gmail\.com|hotmail\.com)', email_extraido, re.IGNORECASE)
                        if match_dominio:
                            pos = match_dominio.start()
                            usuario = email_extraido[:pos]
                            dominio = email_extraido[pos:]
                            email_extraido = f"{usuario}@{dominio}"

                # --- ¡NUEVO BLOQUE DE LIMPIEZA DEL ASUNTO! ---
                # 1. Definimos la frase que queremos eliminar
                frase_a_quitar = "Por ser un asunto de su competencia y a fin de brindar respuesta oportuna,"

                # 2. Limpiamos la frase del asunto que extrajimos
                asunto_limpio = re.sub(frase_a_quitar, '', asunto_extraido, flags=re.IGNORECASE).strip(" ,")
                # --- FIN DEL NUEVO BLOQUE ---

                # 5. Guardar en la sesión (usando el asunto ya limpio)
                request.session['radicado_desde_pdf'] = radicado_extraido
                request.session['asunto_desde_pdf'] = asunto_limpio # <-- ¡AQUÍ ESTÁ EL CAMBIO!
                request.session['peticionario_desde_pdf'] = peticionario_extraido
                request.session['email_desde_pdf'] = email_extraido
                request.session['calidad_peticionario_id_desde_pdf'] = calidad_peticionario_id
                request.session['fecha_recepcion_desde_pdf'] = fecha_para_sesion
                request.session['tipo_tramite_id_desde_pdf'] = tipo_tramite_id 
                return redirect('crear_pqrs')

            except Exception as e:
                 # Si hay un error, borramos el archivo temporal que ya no se usará
                ruta_temporal_a_borrar = os.path.join(settings.MEDIA_ROOT, 'tmp', nombre_archivo_temporal)
                if os.path.exists(ruta_temporal_a_borrar):
                    os.remove(ruta_temporal_a_borrar)

                messages.error(request, f"Error al procesar el PDF: {e}")
                request.session['fecha_recepcion_desde_pdf'] = fecha_recepcion_extraida

                return redirect('crear_pqrs_desde_pdf')
    else:
                form = PdfUploadForm()

    return render(request, 'nucleo/crear_pqrs_desde_pdf.html', {'form': form})


@login_required
@user_passes_test(es_coordinador)
def crear_pqrs(request):
    initial_data = {}
    if 'fecha_recepcion_desde_pdf' in request.session:
        initial_data['fecha_recepcion_inicial'] = request.session.pop('fecha_recepcion_desde_pdf', None)
    if 'radicado_desde_pdf' in request.session: initial_data['radicado'] = request.session.pop('radicado_desde_pdf', None)
    if 'asunto_desde_pdf' in request.session: initial_data['asunto'] = request.session.pop('asunto_desde_pdf', None)
    if 'peticionario_desde_pdf' in request.session: initial_data['peticionario_nombre'] = request.session.pop('peticionario_desde_pdf', None)
    if 'email_desde_pdf' in request.session: initial_data['peticionario_email'] = request.session.pop('email_desde_pdf', None)
    
    if 'calidad_peticionario_id_desde_pdf' in request.session:
        initial_data['calidad_peticionario'] = request.session.pop('calidad_peticionario_id_desde_pdf', None)
    if 'fecha_recepcion_desde_pdf' in request.session:
            initial_data['fecha_recepcion_inicial'] = request.session.pop('fecha_recepcion_desde_pdf', None)
    if 'tipo_tramite_id_desde_pdf' in request.session:
            initial_data['tipo_tramite'] = request.session.pop('tipo_tramite_id_desde_pdf', None)

    if request.method == 'POST':
        form = PqrsForm(request.POST)
        if form.is_valid():
            # --- ¡AQUÍ ESTÁ LA CORRECCIÓN! ---
            nueva_pqrs = form.save() # Guardamos la nueva PQRS en una variable
            
            # --- El resto de tu bloque para adjuntar el PDF ahora funcionará ---
            nombre_archivo_temporal = request.session.pop('pdf_original_temporal', None)
            if nombre_archivo_temporal:
                ruta_temporal = os.path.join(settings.MEDIA_ROOT, 'tmp', nombre_archivo_temporal)
                if os.path.exists(ruta_temporal):
                    with open(ruta_temporal, 'rb') as archivo_pdf:
                        ArchivoAdjunto.objects.create(
                            pqrs=nueva_pqrs, # Ahora esta variable sí existe
                            archivo=File(archivo_pdf, name=nombre_archivo_temporal),
                            descripcion="Documento PDF original de la solicitud.",
                            tipo_archivo='Anexo Peticionario'
                        )
                    os.remove(ruta_temporal)
            
            messages.success(request, 'La PQRS ha sido registrada exitosamente.')
            return redirect('dashboard')
    else:
        form = PqrsForm(initial=initial_data)
        
    return render(request, 'nucleo/pqrs_form.html', {'form': form})

@login_required
def editar_pqrs(request, pqrs_id):
    pqrs = get_object_or_404(Pqrs, pk=pqrs_id)
    es_coord = es_coordinador(request.user)

    if not (pqrs.responsable == request.user or es_coord or request.user.is_superuser):
        return HttpResponseForbidden("No tienes permiso para editar este caso.")

    FormularioAUsar = PqrsForm if (es_coord or request.user.is_superuser) else AbogadoPqrsForm

    if request.method == 'POST':
        form = FormularioAUsar(request.POST, instance=pqrs)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
    else:
        form = FormularioAUsar(instance=pqrs)

    contexto = {
        'form': form,
        'pqrs': pqrs
    }
    return render(request, 'nucleo/pqrs_form.html', contexto)

# nucleo/views.py
# nucleo/views.py
@login_required
def detalle_pqrs(request, pqrs_id):
    pqrs = get_object_or_404(Pqrs, pk=pqrs_id)
    
    # --- Lógica de Permisos (tu código, sin cambios) ---
    es_coord = es_coordinador(request.user)
    user_es_responsable = (pqrs.responsable == request.user)
    
    if not (user_es_responsable or es_coord or request.user.is_superuser):
        return HttpResponseForbidden("No tienes permiso para ver este caso.")

    user_puede_editar = (user_es_responsable or es_coord or request.user.is_superuser)

    # --- Manejo de Formularios (POST) con la nueva lógica de estados ---
    if request.method == 'POST':
        if not (pqrs.responsable == request.user or es_coordinador(request.user) or request.user.is_superuser):
            return HttpResponseForbidden("No tienes permiso para modificar este caso.")

        form_adjuntos = ArchivoAdjuntoForm(request.POST, request.FILES)
        form_seguimiento = SeguimientoForm(request.POST)
        form_respuesta = RespuestaTramiteForm(request.POST, instance=pqrs)

        # 1. Al SUBIR un adjunto
        if 'submit_adjunto' in request.POST and form_adjuntos.is_valid():
            nuevo_adjunto = form_adjuntos.save(commit=False)
            nuevo_adjunto.pqrs = pqrs
            nuevo_adjunto.save()
            
            # --- CAMBIO DE ESTADO ---
            if pqrs.estado == 'Recibido':
                pqrs.estado = 'En Trámite'
                pqrs.save(update_fields=['estado'])
            # --- FIN DEL CAMBIO ---

            messages.success(request, 'Archivo adjuntado correctamente.')
            return redirect('detalle_pqrs', pqrs_id=pqrs.id)
        
        # 2. Al AÑADIR una nota al historial
        elif 'submit_seguimiento' in request.POST and form_seguimiento.is_valid():
            nuevo_seguimiento = form_seguimiento.save(commit=False)
            nuevo_seguimiento.pqrs = pqrs
            nuevo_seguimiento.autor = request.user
            nuevo_seguimiento.save()

            # --- CAMBIO DE ESTADO ---
            if pqrs.estado == 'Recibido':
                pqrs.estado = 'En Trámite'
                pqrs.save(update_fields=['estado'])
            # --- FIN DEL CAMBIO ---

            messages.success(request, 'Se ha añadido una nueva nota al historial.')
            return redirect('detalle_pqrs', pqrs_id=pqrs.id)

        # 3. Al GUARDAR la respuesta definitiva
        elif 'submit_respuesta' in request.POST and form_respuesta.is_valid():
            form_respuesta.save()
            
            # --- CAMBIO DE ESTADO ---
            if pqrs.estado == 'Recibido':
                pqrs.estado = 'En Trámite'
                pqrs.save(update_fields=['estado'])
            # --- FIN DEL CAMBIO ---

            messages.success(request, 'La respuesta definitiva ha sido guardada.')
            return redirect('detalle_pqrs', pqrs_id=pqrs.id)

    # --- El resto de tu función se mantiene exactamente igual ---
    form_adjuntos = ArchivoAdjuntoForm()
    form_seguimiento = SeguimientoForm()
    form_respuesta = RespuestaTramiteForm(instance=pqrs)
    form_traslado = TrasladoPqrsForm()
    adjuntos = pqrs.adjuntos.all()
    seguimientos = pqrs.seguimientos.all()
    
    contexto = {
        'pqrs': pqrs,
        'adjuntos': adjuntos,
        'seguimientos': seguimientos,
        'form_adjuntos': form_adjuntos,
        'form_seguimiento': form_seguimiento,
        'form_respuesta': form_respuesta,
        'form_traslado': form_traslado,
        'puede_gestionar': puede_gestionar_respuesta(request.user),
        'user_puede_editar': user_puede_editar,
        'es_coordinador': es_coord, # <-- ¡Este es el cambio clave!

    }
    
    return render(request, 'nucleo/detalle_pqrs.html', contexto)



@login_required
@user_passes_test(es_coordinador)
def asignar_abogado(request, pqrs_id):
    if request.method == 'POST':
        pqrs = get_object_or_404(Pqrs, pk=pqrs_id)
        abogado_id = request.POST.get('responsable_id')
        
        if abogado_id:
            abogado = get_object_or_404(User, pk=abogado_id)
            pqrs.responsable = abogado
            pqrs.save()
            messages.success(request, f'Se ha asignado a {abogado.get_full_name()} al caso {pqrs.radicado}.')
        else:
            messages.error(request, 'No se seleccionó un abogado válido.')
    return redirect('dashboard')
@login_required
@user_passes_test(es_coordinador_o_abogado)
def enviar_respuesta_email(request, pqrs_id):
    pqrs = get_object_or_404(Pqrs, pk=pqrs_id)

    if request.method == 'POST':
        # --- Lógica de generación de PDF (Tu código original, sin cambios) ---
        def imagen_a_base64(ruta_relativa):
            try:
                with staticfiles_storage.open(ruta_relativa) as image_file:
                    return f"data:image/png;base64,{base64.b64encode(image_file.read()).decode('utf-8')}"
            except Exception as e:
                print(f"ERROR al cargar imagen para PDF: {e}")
                return None
        
        contexto_pdf = {
            'pqrs': pqrs, 'fecha_actual': date.today(),
            'data_uri_encabezado': imagen_a_base64('nucleo/img/encabezado.png'),
            'data_uri_pie': imagen_a_base64('nucleo/img/pie_pagina.png'),
        }
        html_string = render_to_string('nucleo/respuesta_pdf.html', contexto_pdf)
        buffer = BytesIO()
        pisa.CreatePDF(html_string, dest=buffer)
        pdf_en_memoria = buffer.getvalue()
        buffer.close()

        # --- Tu lógica para manejar adjuntos (sin cambios) ---
        adjuntos_para_enviar = []
        ids_adjuntos_seleccionados = request.POST.getlist('adjuntos_a_enviar')
        if ids_adjuntos_seleccionados:
            adjuntos_para_enviar = ArchivoAdjunto.objects.filter(id__in=ids_adjuntos_seleccionados)

        try:
            asunto = f"Respuesta a su solicitud con radicado N° {pqrs.radicado}"
            mensaje = f"Estimado(a) {pqrs.peticionario_nombre},\n\nAdjunto encontrará la respuesta formal a su solicitud y los documentos relacionados.\n\nCordialmente,\n\nVicerrectoría Académica\nUniversidad del Cauca"
            
            email = EmailMessage(asunto, mensaje, settings.EMAIL_HOST_USER, [pqrs.peticionario_email])
            
            nombre_archivo_pdf = f"Respuesta_{pqrs.radicado}.pdf"
            email.attach(nombre_archivo_pdf, pdf_en_memoria, 'application/pdf')

            for adjunto in adjuntos_para_enviar:
                email.attach(adjunto.nombre_corto, adjunto.archivo.read(), None)
            
            email.send()
            
            # --- ¡NUEVA LÓGICA DE CAMBIO DE ESTADO! ---
            # 1. Actualizamos el estado del caso a "Resuelto"
            pqrs.estado = 'Resuelto'
            # 2. Registramos la fecha en que se dio la respuesta
            pqrs.fecha_respuesta = date.today()
            pqrs.save(update_fields=['estado', 'fecha_respuesta'])
            # --- FIN DE LA NUEVA LÓGICA ---
            
            # --- El registro en el historial se mantiene igual ---
            nota_automatica = "Se envió la respuesta por correo electrónico al peticionario."
            nombres_adjuntos = [adj.nombre_corto for adj in adjuntos_para_enviar]
            if nombres_adjuntos:
                nota_automatica += " Archivos adjuntos: " + ", ".join(nombres_adjuntos) + "."

            Seguimiento.objects.create(
                pqrs=pqrs,
                autor=request.user,
                nota=nota_automatica
            )
            
            messages.success(request, f'El correo para la PQRS {pqrs.radicado} ha sido enviado exitosamente.')
        except Exception as e:
            messages.error(request, f'Ocurrió un error al enviar el correo: {e}')
        
        return redirect('detalle_pqrs', pqrs_id=pqrs.id)

    return redirect('detalle_pqrs', pqrs_id=pqrs.id) # Redirige si no es POST

@login_required
@user_passes_test(es_coordinador_o_abogado)
def generar_pdf(request, pqrs_id):
    pqrs = get_object_or_404(Pqrs, pk=pqrs_id)
    
    def imagen_a_base64(ruta_relativa):
        try:
            with staticfiles_storage.open(ruta_relativa) as image_file:
                return f"data:image/png;base64,{base64.b64encode(image_file.read()).decode('utf-8')}"
        except Exception as e:
            print(f"ERROR al codificar la imagen {ruta_relativa}: {e}")
            return None

    contexto = {
        'pqrs': pqrs,
        'fecha_actual': date.today(),
        'data_uri_encabezado': imagen_a_base64('nucleo/img/encabezado.png'),
        'data_uri_pie': imagen_a_base64('nucleo/img/pie_pagina.png'),
    }

    html_string = render_to_string('nucleo/respuesta_pdf.html', contexto)
    buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=buffer)

    if pisa_status.err:
        return HttpResponse('Ocurrió un error al generar el PDF <pre>' + html_string + '</pre>')

    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    nombre_archivo = f"Respuesta_{pqrs.radicado}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response

@login_required
def exportar_excel(request):
    queryset = Pqrs.objects.exclude(estado='Anulado').order_by('-fecha_recepcion_inicial')

    if not es_coordinador(request.user) and not request.user.is_superuser:
        queryset = queryset.filter(responsable=request.user)

    filter_form = PqrsFilterForm(request.GET, user=request.user)
    if filter_form.is_valid():
        q = filter_form.cleaned_data.get('q')
        vigencia = filter_form.cleaned_data.get('vigencia')
        responsable = filter_form.cleaned_data.get('responsable')
        estado = filter_form.cleaned_data.get('estado')
        # ... (lógica de filtros idéntica)
        if q: queryset = queryset.filter(Q(radicado__icontains=q) | Q(asunto__icontains=q))
        if vigencia: queryset = queryset.filter(fecha_recepcion_inicial__year=vigencia)
        elif not estado: queryset = queryset.filter(fecha_recepcion_inicial__year=date.today().year)
        if responsable and (es_coordinador(request.user) or request.user.is_superuser): queryset = queryset.filter(responsable=responsable)
        if estado: queryset = queryset.filter(estado=estado)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte PQRS"

    headers = ["Radicado", "Asunto", "Responsable", "Fecha Recepción", "Fecha Vencimiento", "Fecha Respuesta", "Estado Proceso", "Estado Tiempo", "Peticionario"]
    ws.append(headers)

    for pqrs in queryset:
        responsable_nombre = pqrs.responsable.get_full_name() if pqrs.responsable else "No asignado"
        ws.append([pqrs.radicado, pqrs.asunto, responsable_nombre, pqrs.fecha_recepcion_inicial, pqrs.fecha_vencimiento, pqrs.fecha_respuesta, pqrs.estado, pqrs.estado_tiempo, pqrs.peticionario_nombre])

    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nombre_archivo = f"reporte_pqrs_{date.today().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    wb.save(response)
    return response
@login_required
@user_passes_test(es_coordinador_o_abogado) # Usamos el permiso que ya creamos
def eliminar_adjunto(request, adjunto_id):
    # Buscamos el archivo adjunto específico por su ID
    adjunto = get_object_or_404(ArchivoAdjunto, id=adjunto_id)
    pqrs_id = adjunto.pqrs.id # Guardamos el ID de la PQRS para poder regresar

    if request.method == 'POST':
        # Borra el archivo del almacenamiento
        adjunto.archivo.delete()
        # Borra el registro de la base de datos
        adjunto.delete()
        messages.success(request, 'Archivo adjunto eliminado exitosamente.')
    
    # Redirigimos al usuario de vuelta a la página de detalles de la PQRS
    return redirect('detalle_pqrs', pqrs_id=pqrs_id)
@login_required
@user_passes_test(es_coordinador)
def cerrar_pqrs(request, pqrs_id):
    if request.method == 'POST':
        pqrs = get_object_or_404(Pqrs, pk=pqrs_id)
        if not pqrs.fecha_cierre: # Solo cierra si no está ya cerrado
            pqrs.fecha_cierre = date.today()
            pqrs.save()
            # 2. Crea una nota automática para el historial de seguimiento
            nota_automatica = f"El caso fue cerrado definitivamente por {request.user.get_full_name()}."
            Seguimiento.objects.create(
                pqrs=pqrs,
                autor=request.user,
                nota=nota_automatica
            )
            messages.success(request, f'El caso {pqrs.radicado} ha sido cerrado y archivado exitosamente.')
    return redirect('detalle_pqrs', pqrs_id=pqrs_id)

@login_required
@user_passes_test(es_coordinador)
def reabrir_pqrs(request, pqrs_id):
    if request.method == 'POST':
        pqrs = get_object_or_404(Pqrs, pk=pqrs_id)
        if pqrs.fecha_cierre: # Solo reabre si está cerrado
            # 1. Quita la fecha de cierre para "reabrir" el caso
            pqrs.fecha_cierre = None
            pqrs.save()

            # 2. Crea una nota automática en el historial para la auditoría
            nota_automatica = f"El caso fue reabierto por {request.user.get_full_name()}."
            Seguimiento.objects.create(
                pqrs=pqrs,
                autor=request.user,
                nota=nota_automatica
            )

            messages.success(request, f'El caso {pqrs.radicado} ha sido reabierto exitosamente.')
    return redirect('detalle_pqrs', pqrs_id=pqrs_id)

@login_required
@user_passes_test(es_coordinador)
def trasladar_pqrs(request, pqrs_id):
    pqrs = get_object_or_404(Pqrs, pk=pqrs_id)
    if request.method == 'POST':
        form = TrasladoPqrsForm(request.POST, instance=pqrs)
        if form.is_valid():
            pqrs_actualizada = form.save(commit=False)
            pqrs_actualizada.estado_traslado = 'En Traslado'
            pqrs_actualizada.save()

            # Crea una nota automática en el historial
            nota = f"Caso trasladado a '{pqrs_actualizada.dependencia_trasladada}' por competencia. Queda bajo seguimiento."
            Seguimiento.objects.create(pqrs=pqrs, autor=request.user, nota=nota)

            messages.success(request, f'El caso {pqrs.radicado} ha sido trasladado exitosamente.')
    return redirect('detalle_pqrs', pqrs_id=pqrs_id)

# nucleo/views.py

@login_required
@user_passes_test(es_coordinador)
def deshacer_traslado_pqrs(request, pqrs_id):
    if request.method == 'POST':
        pqrs = get_object_or_404(Pqrs, pk=pqrs_id)
        if pqrs.estado_traslado == 'En Traslado':
            dependencia_anterior = pqrs.dependencia_trasladada
            
            # 1. Revertimos los campos a su estado original
            pqrs.estado_traslado = 'Activo'
            pqrs.dependencia_trasladada = None
            pqrs.save()

            # 2. Creamos una nota automática en el historial para la auditoría
            nota_automatica = f"Se deshizo el traslado a '{dependencia_anterior}'. El caso vuelve a estar activo en la Vicerrectoría."
            Seguimiento.objects.create(pqrs=pqrs, autor=request.user, nota=nota_automatica)
            
            messages.info(request, f'Se ha deshecho el traslado del caso {pqrs.radicado}.')
    return redirect('detalle_pqrs', pqrs_id=pqrs_id)

# 1. Función de chequeo (usa la misma lógica de es_coord que ya tienes)
def es_coordinador_check(user):
    # ASUMIENDO que la función que comprueba el rol se llama 'es_coordinador' o 'es_coord'
    # Ajusta esta línea al nombre real de tu función de utilidad.
    return es_coordinador(user) 

# 2. Vista para manejar la confirmación
@user_passes_test(es_coordinador_check) 
def confirmar_pqrs(request, pqrs_id):
    if request.method == 'POST':
        # Buscamos la PQRS
        pqrs = get_object_or_404(Pqrs, id=pqrs_id)
        
        # 1. Marcar como confirmado
        pqrs.confirmado = True
        
        # 2. Opcional: Asignar responsable si no lo tiene.
        # if not pqrs.responsable:
        #     pqrs.responsable = request.user 
            
        pqrs.save(update_fields=['confirmado'])
        
        messages.success(request, f'PQRS con Radicado {pqrs.radicado} ha sido confirmada y activada.')
        
        # Redirige al detalle
        return redirect('detalle_pqrs', pqrs_id=pqrs.id)
        
    # Si alguien intenta acceder directamente por GET
    return redirect('detalle_pqrs', pqrs_id=pqrs_id)